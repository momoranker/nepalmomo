"""
Nepal Momo Scraper v3 — Expert Edition
========================================
Run anytime to re-scrape Reddit and regenerate the leaderboard.
New mentions and upvotes are automatically picked up on every run.

SCORING FORMULA:
  Score = Σ ( upvotes + 1 ) per unique mention

  • +1  = the recommendation itself has value (even 0-upvote comment)
  • +N  = upvote count of the mentioning comment (community validation)
  • Deduplication: same comment counted once per place, prevents spam inflation

EXTRA SIGNALS:
  • Unique Threads  = how many separate Reddit posts mention this place
  • Trust Tier      = Established / Popular / Known / Emerging
  • Sentiment       = detected from words surrounding the place name
  • Google Maps     = auto-generated search link for each place
"""

import requests, time, re, urllib.parse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADERS    = {"User-Agent": "MomoScraper/3.0 personal Nepal research"}
SUBREDDITS = ["Nepal", "NepalSocial"]

PLACE_REGISTRY = [
    ("Shandaar Momo",             ["shandaar","shandar"],
     "Sankhamul / Baneshwor",     "Shandaar Momo Sankhamul Kathmandu Nepal",          "Buff, Veg"),
    ("Narayan Dai Momo",          ["narayan dai"],
     "Gaushala",                  "Narayan Dai Momo Gaushala Kathmandu Nepal",         "Buff"),
    ("New Food / Dish",           ["new food","new food/dish"],
     "New Road",                  "New Food Momo New Road Kathmandu Nepal",            "Pork"),
    ("Aambo Momo",                ["aambo","ambo momo"],
     "Ason / Pulchowk",           "Aambo Momo Kathmandu Nepal",                       "Buff, Fapar/Kodo"),
    ("Mahabharat Momo",           ["mahabharat momo"],
     "Lagankhel / Patan Dhoka",   "Mahabharat Momo Lagankhel Lalitpur Nepal",         "Buff"),
    ("Oriental Kitchen",          ["oriental kitchen","oriental : baluwatar"],
     "Baluwatar",                 "Oriental Kitchen Baluwatar Kathmandu Nepal",        "Pork"),
    ("Tashi Delek Restaurant",    ["tashi delek","tashi dhelek"],
     "Ekantakuna",                "Tashi Delek Tibetan Restaurant Ekantakuna Nepal",   "Tibetan Buff"),
    ("Magic Momo",                ["magic momo"],
     "Sinamangal / Maharajgunj",  "Magic Momo Sinamangal Kathmandu Nepal",            "Buff, Jhol"),
    ("Everest Momo — Kamaladi",   ["everest momo of kamaladi","everest momo kamaladi","everest momo"],
     "Kamaladi",                  "Everest Momo Kamaladi Kathmandu Nepal",            "Buff"),
    ("Norling Restaurant",        ["norling restaurant","norling"],
     "Kathmandu",                 "Norling Restaurant Kathmandu Nepal",               "Multi"),
    ("Welcome Family Restaurant", ["welcome family restaurant","welcome family"],
     "Shankhamul",                "Welcome Family Restaurant Shankhamul Kathmandu Nepal","Tibetan"),
    ("Moti Didi Ko Momo",         ["moti didi ko momo","moti didi"],
     "Sinamangal",                "Moti Didi Momo Sinamangal Kathmandu Nepal",        "Local"),
    ("Ram Momo — Maru",           ["ram momo at maru","ram momo"],
     "Maru / Kasthamandap",       "Ram Momo Maru Kasthamandap Kathmandu Nepal",       "Buff"),
    ("Bangalamukhi Momo",         ["bangalamukhi mahabharat","bangalamukhi momo","bangalmukhi"],
     "Bangalamukhi",              "Bangalamukhi Momo Kathmandu Nepal",                "Buff"),
    ("Delicious Momo — Jamal",    ["delicious momo,jamal","delicious momo jamal"],
     "Jamal",                     "Delicious Momo Jamal Kathmandu Nepal",             "Buff"),
    ("Mechung",                   ["mechung"],
     "Bouddha",                   "Mechung Restaurant Bouddha Kathmandu Nepal",       "Tibetan"),
    ("Try Again Momo Center",     ["try again momo","try again momo center"],
     "Kathmandu",                 "Try Again Momo Center Kathmandu Nepal",            "Buff"),
    ("Dillibazar Momo",           ["dillibazar ko","dillibazar momo"],
     "Dillibazar",                "Dillibazar Momo Kathmandu Nepal",                  "Buff"),
    ("Bagmati Sweets",            ["bagmati sweets"],
     "Kathmandu",                 "Bagmati Sweets Momo Kathmandu Nepal",              "Veg"),
    ("Pasa Momo",                 ["pasa momo"],
     "Hattiban",                  "Pasa Momo Hattiban Kathmandu Nepal",               "Buff"),
    ("Mint's Hut Cafe",           ["mint's hut","mints hut","mint hut"],
     "Lakeside, Pokhara",         "Mint Hut Cafe Lakeside Pokhara Nepal",             "Multi"),
    ("Himalayan Momo",            ["himalayan momo"],
     "Kathmandu (Chain)",         "Himalayan Momo Nepal",                             "Frozen/Chain"),
    ("Swadisto Momo",             ["swadisto","swadista momo"],
     "Kathmandu",                 "Swadisto Momo Kathmandu Nepal",                    "Buff"),
    ("Haku Dai — Makhan Galli",   ["haku dai","makhan galli"],
     "Makhan Galli / Old Town",   "Haku Dai Momo Makhan Galli Kathmandu Nepal",       "Buff 5-star"),
    ("Bhoso Rahir Momo",          ["bhoso rahir"],
     "Pulchowk",                  "Bhoso Rahir Momo Pulchowk Kathmandu Nepal",        "Buff"),
]

POS_WORDS = {'best','love','favourite','favorite','amazing','great','good','recommend',
             'ramro','mitho','sasto','worth','delicious','perfect','must','awesome',
             'underrated','obsessed','top'}
NEG_WORDS = {'worst','bad','overpriced','overhyped','trash','overrated','hate',
             'disgusting','avoid','terrible','disappointing','namitho','bland',
             'shite','expensive','average','mediocre'}


def get_momo_posts():
    posts = {}
    for sub in SUBREDDITS:
        print(f"\n📋 Getting posts from r/{sub}...")
        for sort in ["top","relevance","new"]:
            for tf in ["all","year"]:
                url = (f"https://www.reddit.com/r/{sub}/search.json"
                       f"?q=momo&sort={sort}&limit=100&restrict_sr=1&t={tf}")
                try:
                    r = requests.get(url, headers=HEADERS, timeout=15)
                    if r.status_code != 200: continue
                    added = 0
                    for p in r.json()["data"]["children"]:
                        d = p["data"]; pid = d.get("id","")
                        if pid not in posts:
                            posts[pid] = {"id":pid,"subreddit":sub,
                                "title":d.get("title",""),"body":d.get("selftext",""),
                                "upvotes":d.get("score",0),"permalink":d.get("permalink",""),
                                "num_comments":d.get("num_comments",0)}
                            added += 1
                    print(f"  ✅ +{added} ({sort}/{tf})")
                    time.sleep(0.8)
                except Exception as e:
                    print(f"  ❌ {e}")
    print(f"\n📦 Total unique posts: {len(posts)}")
    return list(posts.values())


def get_all_comments(posts):
    all_items = []; seen_ids = set()
    for post in posts:
        pid = post["id"]
        if pid not in seen_ids:
            seen_ids.add(pid)
            all_items.append({"type":"post","subreddit":post["subreddit"],
                "text":f"{post['title']} {post['body']}".strip(),
                "upvotes":post["upvotes"],"url":f"https://reddit.com{post['permalink']}",
                "post_title":post["title"]})

    posts_sorted = sorted(posts, key=lambda x: x["num_comments"], reverse=True)
    print(f"\n💬 Fetching comments from {len(posts_sorted)} posts...\n")

    for i, post in enumerate(posts_sorted):
        if not post["permalink"]: continue
        try:
            r = requests.get(f"https://www.reddit.com{post['permalink']}.json?limit=100&depth=3",
                             headers=HEADERS, timeout=15)
            if r.status_code != 200: continue
            data = r.json()
            if len(data) < 2: continue
            cc = [0]
            def extract(clist, sub, purl, ptitle):
                for c in clist:
                    if not isinstance(c, dict): continue
                    cd = c.get("data",{}); body = cd.get("body",""); cid = cd.get("id","")
                    if body and body not in ("[deleted]","[removed]") and len(body)>5 and cid not in seen_ids:
                        seen_ids.add(cid)
                        all_items.append({"type":"comment","subreddit":sub,"text":body,
                            "upvotes":cd.get("score",0),"url":purl,"post_title":ptitle})
                        cc[0] += 1
                    replies = cd.get("replies",{})
                    if isinstance(replies, dict):
                        extract(replies.get("data",{}).get("children",[]), sub, purl, ptitle)
            purl = f"https://reddit.com{post['permalink']}"
            extract(data[1]["data"]["children"], post["subreddit"], purl, post["title"])
            print(f"  [{i+1:3}/{len(posts_sorted)}] '{post['title'][:50]}' → {cc[0]} comments")
            time.sleep(0.6)
        except Exception as e:
            print(f"  ❌ {e}"); time.sleep(1)

    print(f"\n✅ Total: {len(all_items)} items")
    return all_items


def score_places(items):
    url_map = defaultdict(list)
    for item in items:
        url_map[item["url"]].append(item)

    results = []
    for display_name, keywords, area, gmaps_q, momo_type in PLACE_REGISTRY:
        score = total_upvotes = mentions = pos_count = neg_count = 0
        threads = set(); best_quote = ("",0); seen_dedup = set(); best_url = ""

        for url, url_items_list in url_map.items():
            for item in url_items_list:
                ctx_low = item["text"].lower()
                if not any(kw in ctx_low for kw in keywords): continue
                dk = (url, item["text"][:80])
                if dk in seen_dedup: continue
                seen_dedup.add(dk)
                u = max(item["upvotes"], 0)
                score += u + 1; total_upvotes += u; mentions += 1; threads.add(url)
                words = set(ctx_low.split())
                if words & POS_WORDS: pos_count += 1
                if words & NEG_WORDS: neg_count += 1
                if u > best_quote[1]: best_quote = (item["text"][:200], u); best_url = url

        if mentions == 0: continue

        t = len(threads)
        trust = ("⭐ Established" if t>=8 else "🔥 Popular" if t>=5 else
                 "👍 Known" if t>=3 else "📌 Emerging")
        ts = pos_count + neg_count
        sentiment = ("—" if ts==0 else "😍 Loved" if neg_count==0 else
                     "👍 Mostly Positive" if pos_count>neg_count*2 else
                     "⚠️ Debated" if neg_count>pos_count else "🔥 Mixed")

        results.append({"name":display_name,"area":area,"momo_type":momo_type,
            "score":score,"upvotes":total_upvotes,"mentions":mentions,"threads":t,
            "avg_upvote":round(total_upvotes/mentions,1),"trust":trust,"sentiment":sentiment,
            "gmaps_url":f"https://www.google.com/maps/search/{urllib.parse.quote(gmaps_q)}",
            "best_quote":best_quote[0] or "—","best_url":best_url})

    return sorted(results, key=lambda x: -x["score"])


def save_excel(results, filename="nepal_momo_expert_leaderboard.xlsx"):
    wb = Workbook(); ws = wb.active; ws.title = "Momo Leaderboard"
    C_ORANGE="E8651A"; C_DARK_OR="B5420A"; C_ALT="FDEBD0"; C_WHITE="FFFFFF"
    C_DARK="1C1C1C"; C_GREEN="1A7A40"; C_BLUE="1A4E8C"; C_RED="A93226"
    C_GOLD="B8860B"; C_YELLOW="FFFDE7"; C_TEAL="0E6655"

    def side(s='thin',c='DDDDDD'): return Side(style=s,color=c)
    bthin = Border(left=side(),right=side(),top=side(),bottom=side())
    btop  = Border(left=side('medium','E8651A'),right=side('medium','E8651A'),
                   top=side('medium','E8651A'),bottom=side('medium','E8651A'))

    for r_num, (txt, size, bold, bg, h) in enumerate([
        ("🥟  NEPAL MOMO LEADERBOARD  —  Reddit Community Rankings",16,True,"E8651A",40),
        (f"Score=Σ(upvotes+1) per unique mention  •  r/Nepal & r/NepalSocial  •  Updated: {datetime.now().strftime('%d %b %Y')}",9,False,"FFF8F4",18),
        ("Formula: each mention=+1pt + upvote count  |  Unique Threads = separate Reddit posts (higher = more trustworthy)",8,False,"FFFAF7",15),
    ], 1):
        ws.merge_cells(f"A{r_num}:N{r_num}")
        c = ws[f"A{r_num}"]; c.value = txt
        c.font = Font(name="Arial",bold=bold,size=size,color=C_WHITE if r_num==1 else "555555")
        c.fill = PatternFill("solid",fgColor=bg)
        c.alignment = Alignment(horizontal="center",vertical="center",italic=(r_num>1))
        ws.row_dimensions[r_num].height = h

    HEADERS = ["Rank","🏪 Place Name","📍 Area","🗺️ Google\nMaps","🥟 Momo Type",
               "⭐ Score","👍 Total\nUpvotes","💬 Unique\nMentions","🧵 Unique\nThreads",
               "📊 Avg Upvotes\n/ Mention","🏅 Trust Tier","😊 Sentiment",
               "💭 Best Reddit Quote","🔗 Reddit"]
    COL_W = [5,26,24,8,18,8,9,9,9,10,18,18,60,8]
    for col,(h,w) in enumerate(zip(HEADERS,COL_W),1):
        c = ws.cell(row=4,column=col,value=h)
        c.font = Font(name="Arial",bold=True,size=9,color=C_WHITE)
        c.fill = PatternFill("solid",fgColor=C_DARK_OR)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = bthin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 34

    MEDALS = {1:"🥇",2:"🥈",3:"🥉"}
    TC = {"⭐ Established":"145A32","🔥 Popular":"922B21","👍 Known":"1A4E8C","📌 Emerging":"6E2F8B"}

    for i,p in enumerate(results,1):
        row = i+4
        fill = PatternFill("solid",fgColor=C_YELLOW if i<=3 else (C_ALT if i%2==0 else C_WHITE))
        bdr = btop if i<=3 else bthin
        ac = Alignment(horizontal="center",vertical="center")
        al = Alignment(horizontal="left",vertical="center",wrap_text=True)

        for col,val in enumerate([MEDALS.get(i,i),p["name"],p["area"],"Maps ↗",p["momo_type"],
            p["score"],p["upvotes"],p["mentions"],p["threads"],p["avg_upvote"],
            p["trust"],p["sentiment"],p["best_quote"],"View ↗"],1):
            c = ws.cell(row=row,column=col,value=val)
            c.fill=fill; c.border=bdr
            if   col==1:  c.font=Font(name="Arial",size=11,bold=True,color=C_GOLD if i<=3 else C_ORANGE);c.alignment=ac
            elif col==2:  c.font=Font(name="Arial",size=10,bold=(i<=5),color=C_DARK);c.alignment=al
            elif col==3:  c.font=Font(name="Arial",size=9,color="555555");c.alignment=al
            elif col==4:  c.hyperlink=p["gmaps_url"];c.font=Font(name="Arial",size=9,bold=True,color="1558B0",underline="single");c.alignment=ac
            elif col==5:  c.font=Font(name="Arial",size=8,italic=True,color="666666");c.alignment=al
            elif col==6:  c.font=Font(name="Arial",size=12,bold=True,color=C_GREEN);c.alignment=ac
            elif col==7:  c.font=Font(name="Arial",size=10,bold=True,color=C_RED);c.alignment=ac
            elif col==8:  c.font=Font(name="Arial",size=10,bold=True,color=C_BLUE);c.alignment=ac
            elif col==9:  c.font=Font(name="Arial",size=10,bold=True,color=C_TEAL);c.alignment=ac
            elif col==10: c.font=Font(name="Arial",size=9,color=C_DARK);c.number_format="0.0";c.alignment=ac
            elif col==11: c.font=Font(name="Arial",size=9,bold=True,color=TC.get(p["trust"],C_DARK));c.alignment=ac
            elif col==12: c.font=Font(name="Arial",size=9,color=C_DARK);c.alignment=ac
            elif col==13: c.font=Font(name="Arial",size=8,italic=True,color="444444");c.alignment=al
            elif col==14:
                if p["best_url"]: c.hyperlink=p["best_url"]
                c.font=Font(name="Arial",size=9,bold=True,color="1558B0",underline="single");c.alignment=ac
        ws.row_dimensions[row].height=40

    ws.freeze_panes="A5"
    ws.auto_filter.ref=f"A4:N{len(results)+4}"
    wb.save(filename)
    print(f"\n✅ Saved → {filename}")


if __name__ == "__main__":
    print("="*62)
    print("  🥟  Nepal Momo Scraper v3 — Expert Edition")
    print("="*62)
    posts   = get_momo_posts()
    items   = get_all_comments(posts)
    results = score_places(items)

    print(f"\n🏆 Top 15:")
    print(f"{'Rank':<5}{'Place':<30}{'Score':>7}{'Upvotes':>9}{'Mentions':>10}{'Threads':>9}  Trust")
    print("-"*78)
    for i,p in enumerate(results[:15],1):
        print(f"{i:<5}{p['name']:<30}{p['score']:>7}{p['upvotes']:>9}{p['mentions']:>10}{p['threads']:>9}  {p['trust']}")

    save_excel(results)
    print("\n🎉 Done!  Open nepal_momo_expert_leaderboard.xlsx")
    print("💡 Re-run anytime to refresh rankings with latest Reddit data.")

