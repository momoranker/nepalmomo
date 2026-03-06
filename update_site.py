"""
update_site.py — Momo Site Auto-Updater
=========================================
Run this after nepal_momo_scraper_v3.py finishes.
It reads the latest leaderboard Excel and patches
kathmandu_momo_rankings.html with fresh rankings.

Usage:
    python update_site.py

Or let it be called automatically from RUN_ON_WINDOWS.bat
"""

import re, json, sys
from datetime import datetime
from pathlib import Path

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_FILE = "nepal_momo_expert_leaderboard.xlsx"
HTML_FILE  = "kathmandu_momo_rankings.html"
# ────────────────────────────────────────────────────────────────────────────


def read_excel(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌  openpyxl not found. Run:  pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(path)
    ws = wb.active

    # Find the header row (contains "Place Name")
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Place Name" in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("❌  Could not find header row in Excel file.")
        sys.exit(1)

    # Map column names → indices
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            clean = str(cell.value).replace("\n", " ").strip()
            headers[clean] = cell.column - 1  # 0-indexed

    def col(row_vals, *candidates):
        for name in candidates:
            for key, idx in headers.items():
                if name.lower() in key.lower():
                    v = row_vals[idx] if idx < len(row_vals) else None
                    return v
        return None

    places = []
    rank = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = col(row, "Place Name")
        if not name or str(name).startswith("🥇") or str(name).startswith("🥈"):
            continue  # skip medal cells that aren't names

        # Get numeric rank from first column or count
        raw_rank = row[0]
        if isinstance(raw_rank, int):
            rank = raw_rank
        else:
            rank += 1

        score    = col(row, "Score")
        upvotes  = col(row, "Upvote")
        mentions = col(row, "Mention")
        threads  = col(row, "Thread")
        avg_up   = col(row, "Avg")
        area     = col(row, "Area")
        mtype    = col(row, "Momo Type", "Type")
        trust    = col(row, "Trust")
        sentiment= col(row, "Sentiment")
        quote    = col(row, "Quote")

        # Clean trust text (strip emoji prefix)
        trust_clean = str(trust or "").strip()
        for prefix in ["⭐ ", "🔥 ", "👍 ", "📌 "]:
            trust_clean = trust_clean.replace(prefix, "")

        # Clean sentiment
        sentiment_clean = str(sentiment or "—").strip()

        # Safe numeric conversion
        def safe_num(v, default=0):
            try: return float(v) if v is not None else default
            except: return default

        places.append({
            "rank":     rank,
            "name":     str(name).strip(),
            "area":     str(area or "Kathmandu").strip(),
            "type":     str(mtype or "Buff").strip(),
            "score":    int(safe_num(score)),
            "upvotes":  int(safe_num(upvotes)),
            "mentions": int(safe_num(mentions)),
            "threads":  int(safe_num(threads)),
            "avgUp":    round(safe_num(avg_up), 1),
            "trust":    trust_clean,
            "sentiment":sentiment_clean,
            "quote":    str(quote or "—").strip()[:300],
            "gmaps":    f"https://www.google.com/maps/search/{str(name).replace(' ', '+').strip()}+Kathmandu+Nepal",
        })

    return places


def build_js_array(places):
    lines = ["const RAW_DATA = ["]
    for p in places:
        # Escape quotes in strings
        def esc(s): return str(s).replace("\\","\\\\").replace('"','\\"').replace("\n"," ")
        lines.append(
            f'  {{ rank:{p["rank"]}, name:"{esc(p["name"])}", area:"{esc(p["area"])}", '
            f'type:"{esc(p["type"])}", score:{p["score"]}, upvotes:{p["upvotes"]}, '
            f'mentions:{p["mentions"]}, threads:{p["threads"]}, avgUp:{p["avgUp"]}, '
            f'trust:"{esc(p["trust"])}", sentiment:"{esc(p["sentiment"])}", '
            f'quote:"{esc(p["quote"])}", '
            f'gmaps:"https://www.google.com/maps/search/{p["name"].replace(chr(32),"+").replace(chr(34),"")}+Kathmandu+Nepal" }},'
        )
    lines.append("];")
    return "\n".join(lines)


def update_html(html_path, new_js_array, places, updated_label):
    html = Path(html_path).read_text(encoding="utf-8")

    # Replace RAW_DATA block
    pattern = r"const RAW_DATA = \[[\s\S]*?\];"
    if not re.search(pattern, html):
        print("❌  Could not find RAW_DATA block in HTML. Make sure you're using the right HTML file.")
        sys.exit(1)

    html = re.sub(pattern, new_js_array, html)

    # Update "Last Updated" date in stats bar
    html = re.sub(
        r"document\.getElementById\('statUpdated'\)\.textContent\s*=\s*'[^']*';",
        f"document.getElementById('statUpdated').textContent = '{updated_label}';",
        html
    )

    # Update total places / mentions in stats (optional dynamic calc already in JS)
    Path(html_path).write_text(html, encoding="utf-8")


def main():
    print("=" * 55)
    print("  🥟  Momo Site Updater")
    print("=" * 55)

    # Check files exist
    excel_path = Path(EXCEL_FILE)
    html_path  = Path(HTML_FILE)

    if not excel_path.exists():
        print(f"\n❌  Excel file not found: {EXCEL_FILE}")
        print("    Run nepal_momo_scraper_v3.py first!")
        sys.exit(1)

    if not html_path.exists():
        print(f"\n❌  HTML file not found: {HTML_FILE}")
        print("    Make sure kathmandu_momo_rankings.html is in the same folder.")
        sys.exit(1)

    print(f"\n📖  Reading {EXCEL_FILE}...")
    places = read_excel(excel_path)
    print(f"    ✅  Found {len(places)} momo spots")

    if not places:
        print("❌  No data found in Excel. Aborting.")
        sys.exit(1)

    print(f"\n🔧  Building new rankings data...")
    js_array = build_js_array(places)

    now   = datetime.now()
    label = now.strftime("%b %Y")

    print(f"\n✏️   Patching {HTML_FILE}...")
    update_html(str(html_path), js_array, places, label)

    print(f"\n🎉  Done! Site updated with {len(places)} spots.")
    print(f"    Top 5 right now:")
    for p in places[:5]:
        print(f"      #{p['rank']}  {p['name']:<30}  Score: {p['score']}")
    print(f"\n    Open {HTML_FILE} in your browser to see the update.\n")


if __name__ == "__main__":
    main()
